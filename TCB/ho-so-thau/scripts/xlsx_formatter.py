#!/usr/bin/env python3
"""
Format .xlsx files to match standards and apply review tracking.

This script formats Excel files according to formatting standards and optionally
applies review tracking with audit columns, color coding, and change logs.

Usage:
    python scripts/xlsx_formatter.py output/pricing.xlsx
    python scripts/xlsx_formatter.py output/pricing.xlsx --no-review-mode
    python scripts/xlsx_formatter.py output/pricing.xlsx --config config.yaml --changes changes.json
"""

import argparse
import json
import logging
import sys
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.styles.differential import DifferentialStyle
    from openpyxl.formatting.rule import CellIsRule, FormulaRule
    from openpyxl.comments import Comment
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

try:
    import yaml
except ImportError:
    yaml = None
    logging.warning("PyYAML not available. Config file support will be limited.")


logging.basicConfig(
    level=logging.INFO,
    format='%(levelname)s: %(message)s'
)
logger = logging.getLogger(__name__)


# Color constants
HEADER_BG = "1F4E79"  # Dark blue
HEADER_TEXT = "FFFFFF"  # White
NEW_ROW_COLOR = "E2EFDA"  # Light green
MODIFIED_ROW_COLOR = "FFF2CC"  # Yellow
NEGATIVE_COLOR = "FF0000"  # Red
POSITIVE_COLOR = "00FF00"  # Green

# Audit column names
AUDIT_COLUMNS = ["_Status", "_Change_Desc", "_Updated_By"]


def load_config(config_path: Optional[Path]) -> Dict[str, Any]:
    """
    Load configuration from YAML file.
    
    Args:
        config_path: Path to config.yaml file
        
    Returns:
        Dictionary with config values, or empty dict if not found
    """
    if config_path is None:
        # Try default location
        config_path = Path("config.yaml")
    
    if not config_path.exists():
        logger.warning(f"Config file not found: {config_path}")
        return {}
    
    if yaml is None:
        logger.warning("PyYAML not installed. Cannot read config file.")
        return {}
    
    try:
        with open(config_path, 'r', encoding='utf-8') as f:
            config = yaml.safe_load(f) or {}
        return config
    except Exception as e:
        logger.error(f"Error reading config file: {e}")
        return {}


def load_changes(changes_path: Optional[Path]) -> List[Dict[str, Any]]:
    """
    Load changes from JSON file.
    
    Args:
        changes_path: Path to changes.json file
        
    Returns:
        List of change dictionaries
    """
    if changes_path is None:
        logger.warning("No changes.json provided. Review tracking will be minimal.")
        return []
    
    if not changes_path.exists():
        logger.warning(f"Changes file not found: {changes_path}")
        return []
    
    try:
        with open(changes_path, 'r', encoding='utf-8') as f:
            changes = json.load(f)
        if not isinstance(changes, list):
            logger.error("changes.json must contain a list of change objects")
            return []
        return changes
    except json.JSONDecodeError as e:
        logger.error(f"Invalid JSON in changes file: {e}")
        return []
    except Exception as e:
        logger.error(f"Error reading changes file: {e}")
        return []


def get_author_info(config: Dict[str, Any]) -> str:
    """
    Extract author name from config.
    
    Args:
        config: Configuration dictionary
        
    Returns:
        Author name string
    """
    author_name = config.get("author", {}).get("name", "Unknown")
    return author_name


def format_header_row(ws: openpyxl.worksheet.worksheet.Worksheet, header_row: int = 1):
    """
    Format the header row according to standards.
    
    Args:
        ws: Worksheet to format
        header_row: Row number for header (default: 1)
    """
    header_fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color=HEADER_TEXT)
    
    for cell in ws[header_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Freeze panes (freeze top row)
    ws.freeze_panes = f"A{header_row + 1}"
    
    # Enable auto-filter
    if ws.max_row > header_row:
        ws.auto_filter.ref = ws[f"A{header_row}:{get_column_letter(ws.max_column)}{header_row}"]


def auto_fit_columns(ws: openpyxl.worksheet.worksheet.Worksheet, min_width: int = 12):
    """
    Auto-fit column widths with minimum width.
    
    Args:
        ws: Worksheet to format
        min_width: Minimum column width in characters
    """
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if cell.value:
                    # Calculate length considering formatting
                    cell_value = str(cell.value)
                    max_length = max(max_length, len(cell_value))
            except:
                pass
        
        adjusted_width = max(max_length + 2, min_width)
        ws.column_dimensions[column_letter].width = adjusted_width


def apply_data_formatting(ws: openpyxl.worksheet.worksheet.Worksheet, 
                         data_start_row: int = 2):
    """
    Apply standard data formatting to all data rows.
    
    Args:
        ws: Worksheet to format
        data_start_row: First row of data (after header)
    """
    data_font = Font(name="Calibri", size=11)
    
    for row in ws.iter_rows(min_row=data_start_row, max_row=ws.max_row):
        for cell in row:
            # Skip audit columns (will be formatted separately)
            if cell.column > ws.max_column - len(AUDIT_COLUMNS):
                continue
            
            # Apply font
            if cell.font.name != "Calibri" or cell.font.size != 11:
                cell.font = data_font
            
            # Detect and format numbers/currency
            if isinstance(cell.value, (int, float)):
                # Check if it's a currency value (heuristic: if column header contains price/cost/total)
                header_value = str(ws.cell(row=1, column=cell.column).value or "").lower()
                if any(keyword in header_value for keyword in ["price", "cost", "total", "amount", "gia", "tong"]):
                    cell.number_format = '#,##0 "VND"'
                elif isinstance(cell.value, float) and 0 <= cell.value <= 1:
                    # Likely a percentage
                    cell.number_format = '0.0%'
                else:
                    cell.number_format = '#,##0'
            
            # Format dates (if cell contains date)
            if isinstance(cell.value, datetime):
                cell.number_format = 'dd/mm/yyyy'
            
            # Negative numbers in red
            if isinstance(cell.value, (int, float)) and cell.value < 0:
                cell.font = Font(name="Calibri", size=11, color=NEGATIVE_COLOR)


def apply_conditional_formatting(ws: openpyxl.worksheet.worksheet.Worksheet,
                                data_start_row: int = 2):
    """
    Apply conditional formatting rules.
    
    Args:
        ws: Worksheet to format
        data_start_row: First row of data
    """
    if ws.max_row < data_start_row:
        return  # No data rows
    
    # Red background for negative values
    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    negative_rule = CellIsRule(operator="lessThan", formula=["0"], fill=red_fill)
    
    # Green background for positive targets (heuristic: values > 1000)
    green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    positive_rule = CellIsRule(operator="greaterThan", formula=["1000"], fill=green_fill)
    
    # Apply to numeric columns (skip header and audit columns)
    # Calculate last data column (before audit columns)
    last_data_col = ws.max_column - len(AUDIT_COLUMNS) if ws.max_column > len(AUDIT_COLUMNS) else ws.max_column
    
    for col_idx in range(1, last_data_col + 1):
        col_letter = get_column_letter(col_idx)
        range_str = f"{col_letter}{data_start_row}:{col_letter}{ws.max_row}"
        
        try:
            ws.conditional_formatting.add(range_str, negative_rule)
            ws.conditional_formatting.add(range_str, positive_rule)
        except Exception as e:
            logger.debug(f"Could not apply conditional formatting to {range_str}: {e}")


def validate_formulas(ws: openpyxl.worksheet.worksheet.Worksheet) -> List[str]:
    """
    Validate formulas and check for errors.
    
    Args:
        ws: Worksheet to validate
        
    Returns:
        List of error messages
    """
    errors = []
    error_patterns = ["#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#NUM!", "#NULL!"]
    
    for row in ws.iter_rows():
        for cell in row:
            if cell.data_type == 'f':  # Formula cell
                try:
                    # Evaluate formula to check for errors
                    if cell.value:
                        formula_str = str(cell.value)
                        for pattern in error_patterns:
                            if pattern in formula_str:
                                errors.append(
                                    f"Formula error in {cell.coordinate}: {formula_str}"
                                )
                except Exception as e:
                    errors.append(
                        f"Error evaluating formula in {cell.coordinate}: {e}"
                    )
    
    return errors


def set_print_area(ws: openpyxl.worksheet.worksheet.Worksheet):
    """
    Set print area for all data ranges.
    
    Args:
        ws: Worksheet to configure
    """
    if ws.max_row > 0 and ws.max_column > 0:
        print_area = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
        ws.print_area = print_area


def add_audit_columns(ws: openpyxl.worksheet.worksheet.Worksheet, 
                     changes: List[Dict[str, Any]],
                     author: str,
                     data_start_row: int = 2) -> Dict[int, Dict[str, Any]]:
    """
    Add audit columns to the rightmost of the worksheet.
    
    Args:
        ws: Worksheet to modify
        changes: List of change dictionaries
        author: Author name
        data_start_row: First row of data
        
    Returns:
        Dictionary mapping row numbers to change info
    """
    # Find the rightmost column before audit columns
    last_data_col = ws.max_column
    
    # Add audit column headers
    header_row = data_start_row - 1
    for idx, col_name in enumerate(AUDIT_COLUMNS):
        col_idx = last_data_col + idx + 1
        cell = ws.cell(row=header_row, column=col_idx)
        cell.value = col_name
        cell.fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
        cell.font = Font(name="Calibri", size=11, bold=True, color=HEADER_TEXT)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Map changes by row number
    changes_by_row: Dict[int, Dict[str, Any]] = {}
    for change in changes:
        if change.get("sheet") == ws.title:
            row_num = change.get("row")
            if row_num:
                changes_by_row[row_num] = change
    
    # Apply audit data to rows
    for row_num in range(data_start_row, ws.max_row + 1):
        change = changes_by_row.get(row_num, {})
        change_type = change.get("change_type", "UNCHANGED")
        description = change.get("description", "")
        old_value = change.get("old_value")
        new_value = change.get("new_value")
        change_author = change.get("author", author)
        change_model = change.get("model", "")
        change_date = change.get("date", datetime.now().strftime("%Y-%m-%d"))
        
        # Status column
        status_cell = ws.cell(row=row_num, column=last_data_col + 1)
        status_cell.value = change_type
        
        # Change description column
        desc_cell = ws.cell(row=row_num, column=last_data_col + 2)
        desc_cell.value = description
        
        # Updated by column
        updated_cell = ws.cell(row=row_num, column=last_data_col + 3)
        updated_cell.value = f"{change_author} / {change_model} / {change_date}"
        
        # Apply row color based on change type (only to data cells, not header)
        if row_num >= data_start_row:
            row_fill = None
            if change_type == "NEW":
                row_fill = PatternFill(start_color=NEW_ROW_COLOR, end_color=NEW_ROW_COLOR, fill_type="solid")
            elif change_type == "MODIFIED":
                row_fill = PatternFill(start_color=MODIFIED_ROW_COLOR, end_color=MODIFIED_ROW_COLOR, fill_type="solid")
            
            if row_fill:
                # Apply to data cells only (skip audit columns which will be formatted separately)
                for col_idx in range(1, last_data_col + 1):
                    cell = ws.cell(row=row_num, column=col_idx)
                    # Only apply if cell doesn't have a custom fill (preserve header formatting)
                    if cell.fill.fill_type is None or cell.fill.start_color.index == "00000000":
                        cell.fill = row_fill
    
    return changes_by_row


def add_cell_comments(ws: openpyxl.worksheet.worksheet.Worksheet,
                     changes: List[Dict[str, Any]],
                     author: str):
    """
    Add cell comments for modified cells.
    
    Args:
        ws: Worksheet to modify
        changes: List of change dictionaries
        author: Author name
    """
    for change in changes:
        if change.get("sheet") != ws.title:
            continue
        
        row = change.get("row")
        col = change.get("column")  # Optional: specific column
        old_value = change.get("old_value")
        new_value = change.get("new_value")
        description = change.get("description", "")
        change_author = change.get("author", author)
        change_model = change.get("model", "")
        
        if row:
            # If column specified, use it; otherwise, find the cell with new_value
            if col:
                cell = ws.cell(row=row, column=col)
            else:
                # Find cell containing new_value
                cell = None
                for c in ws[row]:
                    if c.value == new_value:
                        cell = c
                        break
                
                if cell is None:
                    # Use first data cell in row
                    cell = ws.cell(row=row, column=1)
            
            comment_text = f"{change_author}/{change_model} Changed from '{old_value}' to '{new_value}' - {description}"
            cell.comment = Comment(comment_text, change_author)


def create_change_log_sheet(wb: openpyxl.workbook.workbook.Workbook,
                            changes: List[Dict[str, Any]],
                            author: str) -> openpyxl.worksheet.worksheet.Worksheet:
    """
    Create a _Change_Log sheet with all changes.
    
    Args:
        wb: Workbook to add sheet to
        changes: List of change dictionaries
        author: Author name
        
    Returns:
        The created worksheet
    """
    # Remove existing _Change_Log sheet if present
    if "_Change_Log" in wb.sheetnames:
        wb.remove(wb["_Change_Log"])
    
    ws = wb.create_sheet("_Change_Log")
    
    # Headers
    headers = ["Sheet", "Cell/Row", "Change Type", "Old Value", "New Value", 
               "Description", "Author", "Date"]
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=idx)
        cell.value = header
        cell.font = Font(name="Calibri", size=11, bold=True)
        cell.fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
        cell.font = Font(name="Calibri", size=11, bold=True, color=HEADER_TEXT)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Summary row
    new_count = sum(1 for c in changes if c.get("change_type") == "NEW")
    modified_count = sum(1 for c in changes if c.get("change_type") == "MODIFIED")
    formula_count = sum(1 for c in changes if "formula" in str(c.get("description", "")).lower())
    
    summary_text = f"Total: {new_count} new rows, {modified_count} modified cells, {formula_count} formula updates"
    ws.cell(row=2, column=1).value = summary_text
    ws.cell(row=2, column=1).font = Font(name="Calibri", size=11, bold=True, italic=True)
    
    # Data rows
    for idx, change in enumerate(changes, start=3):
        ws.cell(row=idx, column=1).value = change.get("sheet", "")
        ws.cell(row=idx, column=2).value = f"Row {change.get('row', '?')}"
        ws.cell(row=idx, column=3).value = change.get("change_type", "")
        ws.cell(row=idx, column=4).value = change.get("old_value", "")
        ws.cell(row=idx, column=5).value = change.get("new_value", "")
        ws.cell(row=idx, column=6).value = change.get("description", "")
        ws.cell(row=idx, column=7).value = change.get("author", author)
        ws.cell(row=idx, column=8).value = change.get("date", "")
    
    # Format the sheet
    format_header_row(ws, header_row=1)
    auto_fit_columns(ws)
    
    return ws


def set_workbook_metadata(wb: openpyxl.workbook.workbook.Workbook,
                         author: str):
    """
    Set workbook metadata (author, etc.).
    
    Args:
        wb: Workbook to modify
        author: Author name
    """
    wb.properties.creator = author
    wb.properties.lastModifiedBy = author
    wb.properties.modified = datetime.now()


def format_worksheet(ws: openpyxl.worksheet.worksheet.Worksheet,
                    changes: List[Dict[str, Any]],
                    author: str,
                    review_mode: bool = True,
                    data_start_row: int = 2):
    """
    Format a single worksheet according to standards.
    
    Args:
        ws: Worksheet to format
        changes: List of change dictionaries
        author: Author name
        review_mode: Whether to apply review tracking
        data_start_row: First row of data (default: 2, assuming row 1 is header)
    """
    logger.info(f"Formatting worksheet: {ws.title}")
    
    # Format header row
    format_header_row(ws, header_row=data_start_row - 1)
    
    # Apply data formatting
    apply_data_formatting(ws, data_start_row=data_start_row)
    
    # Apply conditional formatting
    apply_conditional_formatting(ws, data_start_row=data_start_row)
    
    # Auto-fit columns
    auto_fit_columns(ws)
    
    # Set print area
    set_print_area(ws)
    
    # Review tracking
    if review_mode:
        # Add audit columns
        add_audit_columns(ws, changes, author, data_start_row=data_start_row)
        
        # Add cell comments
        add_cell_comments(ws, changes, author)
    
    # Validate formulas
    formula_errors = validate_formulas(ws)
    if formula_errors:
        logger.warning(f"Formula errors found in {ws.title}:")
        for error in formula_errors:
            logger.warning(f"  - {error}")


def format_workbook(file_path: Path,
                   config_path: Optional[Path] = None,
                   changes_path: Optional[Path] = None,
                   review_mode: bool = True) -> bool:
    """
    Format an Excel workbook according to standards.
    
    Args:
        file_path: Path to .xlsx file
        config_path: Path to config.yaml (optional)
        changes_path: Path to changes.json (optional)
        review_mode: Whether to apply review tracking
        
    Returns:
        True if successful, False otherwise
    """
    if not file_path.exists():
        logger.error(f"File not found: {file_path}")
        return False
    
    if not file_path.suffix.lower() == '.xlsx':
        logger.error(f"File must be .xlsx format: {file_path}")
        return False
    
    # Load config and changes
    config = load_config(config_path)
    author = get_author_info(config)
    changes = load_changes(changes_path) if review_mode else []
    
    logger.info(f"Formatting workbook: {file_path}")
    logger.info(f"Author: {author}")
    logger.info(f"Review mode: {review_mode}")
    logger.info(f"Changes loaded: {len(changes)}")
    
    try:
        # Open workbook
        wb = openpyxl.load_workbook(file_path)
        
        # Set workbook metadata
        set_workbook_metadata(wb, author)
        
        # Format each worksheet
        for ws in wb.worksheets:
            # Skip _Change_Log sheet (will be created/recreated)
            if ws.title == "_Change_Log":
                continue
            
            # Determine data start row (assume row 1 is header, or find first non-empty row)
            data_start_row = 2
            if ws.max_row > 0:
                # Check if row 1 has headers
                first_row_has_data = any(cell.value for cell in ws[1])
                if not first_row_has_data and ws.max_row > 1:
                    # Try row 2 as header
                    data_start_row = 3
            
            format_worksheet(ws, changes, author, review_mode, data_start_row=data_start_row)
        
        # Create change log sheet if review mode
        if review_mode and changes:
            create_change_log_sheet(wb, changes, author)
        
        # Save workbook
        wb.save(file_path)
        logger.info(f"Successfully formatted: {file_path}")
        return True
        
    except Exception as e:
        logger.error(f"Error formatting workbook: {e}", exc_info=True)
        return False


def main():
    """Main entry point."""
    parser = argparse.ArgumentParser(
        description="Format .xlsx files to match standards and apply review tracking",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python scripts/xlsx_formatter.py output/pricing.xlsx
  python scripts/xlsx_formatter.py output/pricing.xlsx --no-review-mode
  python scripts/xlsx_formatter.py output/pricing.xlsx --config config.yaml --changes changes.json
        """
    )
    
    parser.add_argument(
        "file",
        type=Path,
        help="Path to .xlsx file to format"
    )
    
    parser.add_argument(
        "--review-mode",
        action="store_true",
        default=True,
        help="Enable review tracking (default: enabled)"
    )
    
    parser.add_argument(
        "--no-review-mode",
        dest="review_mode",
        action="store_false",
        help="Disable review tracking"
    )
    
    parser.add_argument(
        "--config",
        type=Path,
        default=None,
        help="Path to config.yaml file (default: config.yaml)"
    )
    
    parser.add_argument(
        "--changes",
        type=Path,
        default=None,
        help="Path to changes.json file"
    )
    
    parser.add_argument(
        "-v", "--verbose",
        action="store_true",
        help="Enable verbose logging"
    )
    
    args = parser.parse_args()
    
    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)
    
    success = format_workbook(
        file_path=args.file,
        config_path=args.config,
        changes_path=args.changes,
        review_mode=args.review_mode
    )
    
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
